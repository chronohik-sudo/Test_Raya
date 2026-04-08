from flask import Flask, render_template, request, redirect
from openpyxl import Workbook, load_workbook
import os
from openpyxl import Workbook, load_workbook

FILE_NAME = "responses.xlsx"

def save_to_excel(data):
    if not os.path.exists(FILE_NAME):
        wb = Workbook()
        ws = wb.active
        ws.append(["First Name", "Last Name", "Email", "Company", "Attendance"])
    else:
        wb = load_workbook(FILE_NAME)
        ws = wb.active

    ws.append(data)
    wb.save(FILE_NAME)

@app.route('/')
def form():
    return render_template('form.html')

@app.route('/submit', methods=['POST'])
def submit():
    fname = request.form['fname']
    lname = request.form['lname']
    email = request.form['email']
    company = request.form['company']
    attendance = request.form['attendance']

        # ✅ THIS IS WHERE YOU PUT IT
    save_to_excel([fname, lname, email, company, attendance])

    return redirect('/')

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=10000)
